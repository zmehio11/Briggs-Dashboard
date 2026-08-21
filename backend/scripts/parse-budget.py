#!/usr/bin/env python3
"""
Reads the "Monthly Budget" tab of the Briggs operating budget workbook and
writes backend/prisma/budget-seed.json — one row per calendar month, with
fiscal "Month 1" (column B, labeled Jan) mapped onto whichever real month
the restaurant opened.

Usage:
  python3 parse-budget.py <path-to-xlsx> <opening-year> <opening-month>

Example (restaurant opened June 2026):
  python3 parse-budget.py ~/Downloads/Briggs_Operating_Budget.xlsx 2026 6
"""
import json
import sys

import openpyxl

MONTH_COLUMNS = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]  # Jan..Dec columns
ROW_TOTAL_REVENUE = 8
ROW_TOTAL_COGS = 13
ROW_TOTAL_LABOR = 22


def main():
    if len(sys.argv) != 4:
        print(__doc__)
        sys.exit(1)

    xlsx_path, opening_year, opening_month = sys.argv[1], int(sys.argv[2]), int(sys.argv[3])

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Monthly Budget"]

    rows = []
    year, month = opening_year, opening_month
    for col in MONTH_COLUMNS:
        rows.append(
            {
                "year": year,
                "month": month,
                "totalRevenue": ws[f"{col}{ROW_TOTAL_REVENUE}"].value,
                "totalCogs": ws[f"{col}{ROW_TOTAL_COGS}"].value,
                "totalLabor": ws[f"{col}{ROW_TOTAL_LABOR}"].value,
            }
        )
        month += 1
        if month > 12:
            month = 1
            year += 1

    out_path = "prisma/budget-seed.json"
    with open(out_path, "w") as f:
        json.dump(rows, f, indent=2)

    print(f"Wrote {len(rows)} months to {out_path}:")
    for r in rows:
        print(f"  {r['year']}-{r['month']:02d}: revenue={r['totalRevenue']} cogs={r['totalCogs']} labor={r['totalLabor']}")


if __name__ == "__main__":
    main()
